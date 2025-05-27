import os
import openpyxl as pyxl


# 변경할 폴더가 들어있는 상위 메인 폴더 경로 지정

# dirName = 'D:/20250527_재무팀요청/2024년/2024년'
dirName = input("대조할 폴더들이 포함되어 있는 경로를 입력하세요 : ")
path = os.listdir(dirName)

# 비교할 엑셀 파일 경로 지정
# excelPath = "D:/20250527_재무팀요청/외환검사자료제출양식_(주)코스메카코리아_250526.xlsx"
excelPath = input("엑셀 파일의 경로를 입력하세요(.xlsx까지) : ")

renamedFolder = []
invoiceNumList = {}

# 엑셀 파일 OPEN, 비교할 SHEET 지정
wb = pyxl.load_workbook(excelPath)
sheet = wb.active
value = sheet.cell(3,2).value
sheetRows = sheet.max_row


# 실제 폴더명에서 하이픈 제거
def replaceHyphen() :
    for names in path:
        rep = names.replace("-", "")
        renamedFolder.append(rep)


# 엑셀 행별 Dictionary Type 생성
# EX) {연번 : 수출신고번호}
def matchExcelRows() :
    for i in range(3, sheetRows) :
        yearNumber = sheet.cell(i,1).value
        invoiceNumber = sheet.cell(i,2).value

        invoiceNumList[yearNumber] = invoiceNumber


# 폴더별 수출번호 & Dictionary 비교
# 수출번호 폴더 수정 → 연번-수출번호 형식으로 변경
def matchFolderNExcel() :
    # numListLength = len(invoiceNumList)
    for folderName in path :
        repName = folderName.replace("-", "")
        for key, value in invoiceNumList.items() :
            if(repName == value):
                os.rename(dirName + '/' + folderName, f"{dirName}/{key}-{folderName}")
                



# replaceHyphen()
matchExcelRows()
matchFolderNExcel()

# print("\nKEYS() : ", invoiceNumList[1])
print("===== TARGET FOLDER LIST =====\n", path)
print("\nMAX OF SHEET ROWS : ", sheetRows)
print("\nLENGTH OF DICTIONARY : ", len(invoiceNumList))
print("\nVALUES() : ", invoiceNumList[1])
print("\nFOLDER PATH : ", path[0])


# For Debugging
# for folderName in path :
#     repName = folderName.replace("-", "")
#     for key, value in invoiceNumList.items() :
#         if(repName == value) :
#             print("repName == value")
#         else :
#             print("repName doesn't same as value")



# print("\n== LENGTH OF PATH ==\n", len(path))
# print("\n== LIST ==\n", list, end=", ")
# print("\n*** EXCEL DICTIONARY ***\n", invoiceNumList, end=", ")