import os
import sys
import threading
import openpyxl as pyxl
import tkinter as tk
import time
import pyfiglet
import keyboard
import msvcrt

from tkinter import filedialog

class FolderRenamer :
        
    def __init__(self) :
        self.running = True


    # ESC 키 눌릴시 프로그램 종료
    def on_key_exit(self) :
        keyboard.wait('esc')
        print("ESC가 눌렸습니다. 프로그램을 종료합니다")
        self.running = False
        sys.exit(0)

    # 최초 프로그램 시작 전 사용자 확인
    def waitingForUser(self) :
        print("프로그램을 시작하려면 아무키나 눌러주세요.")
        keyboard.read_event()


    # 프로그램 타이틀 출력
    def on_key_title(self, ver, author, date) :
        print("\n===================================================================================")
        print(f"Author : {author}\nCreated Date : {date}")
        title = "CosmeccaKorea FolderRenamer"
        ascii_art = pyfiglet.figlet_format(title)
        print(f"{ascii_art.rstrip()}     {str(ver)}")
        print("===================================================================================")
        print("")


    # 시트 추출시 로딩 애니메이션 + 스레드 객체 인자
    def loading_animation(self, stop_event):
        spinner = ['|', '/', '-', '\\']
        idx = 0
        print(f"\nLoading...... ", end="", flush=True)
        while not stop_event.is_set():
            print(f"\b{spinner[idx % len(spinner)]}", end="", flush=True)
            idx += 1
            time.sleep(0.1)
        print("\bDone!", flush=True)


    # 실제 대상이 될 폴더와 Dictionary 비교 후 폴더명 변경
    def matchNRename(self, chosenSheet, sheet, maxRow):
        print("\n== 변경이 필요한 폴더의 상위 경로를 선택해주세요 ==")
        print("예) 수입(원료)관련자료-2020년 일 경우 \"수입(원료)관련자료\" 폴더 선택")
        dirPath = filedialog.askdirectory()
        splitDir = dirPath.split("/")[-1]
        print(f"\n- 선택된 폴더 : {splitDir}")

        # 사용자 데이터 기입 기반 폴더명 변경
        while True:
            # Dictionary 에 Key{연번}, Value{수입신고번호} Append
            # 첫번째 데이터가 3번행이기에 for 문의 range를 3으로 잡음
            # 이후 데이터가 다른 행에 있다면 해당 range 변경해야함
            startRow = int(input("\n- 데이터가 시작되는 행을 기입해주세요 : "))
            startCol = int(input("- 첫번째 데이터가 시작되는 열을 기입해주세요(연번) : "))
            startCol2 = int(input("- 두번째 데이터가 시작되는 열을 기입해주세요(신고번호) : "))

            print(f"\n1) 데이터가 시작되는 행 : {startRow}\n2) 첫번째 데이터가 시작되는 열(연번) : {startCol}\n3) 두번째 데이터가 시작되는 열(신고번호) : {startCol2}")

            print("\n\n\n          ※ 주 의 ※")
            confirm = str(input("\n다음 단계는 실제로 폴더명 변경이 이루어집니다.\n경로와 행, 열 정보를 정확히 기입했습니까?(yes/no) : "))

            if(confirm.lower() == "yes"):

                numList = {}
                for i in range(startRow, maxRow):  
                    yearNumber = sheet.cell(i,startCol).value
                    invoiceNumber = sheet.cell(i,startCol2).value

                    numList[yearNumber] = invoiceNumber

                realDir = os.listdir(dirPath)
                # print(realDir)

                # realDir(연도별을 나타내는 폴더들)
                # folders(연도별 하위 폴더들, 실제 폴더명이 변경될 폴더들)
                for i in realDir:
                    folders = os.listdir(dirPath + "/" + i)
                    # print(i)
                    # print(folders)
                    for k in folders:
                        repName = k.replace("-", "")
                        # print(k)
                        # print(repName)
                        for key, value in numList.items() :
                            if (repName == value):
                                os.rename(dirPath + "/" + i + "/" + k, f"{dirPath}/{i}/{chosenSheet}-{key}")

                # [로딩이벤트] 1. 로딩 이벤트를 위한 스레드 객체 생성
                stop_event = threading.Event()

                # [로딩이벤트] 2.로딩 이벤트 중 실행될 스레드
                loading_thread = threading.Thread(target=self.loading_animation, args=(stop_event,))
                loading_thread.start()

                # [로딩이벤트] 3.로딩 이벤트 종료
                stop_event.set()
                loading_thread.join()

                break

            elif(confirm.lower() == "no"):
                print("데이터를 다시 기입하세요")

            else : 
                print("올바르지 않은 선택입니다.")
            


    # 메인 함수
    def on_key_start(self) :
        # tkinter import 및 withdraw 활용하여 창 숨기기
        root = tk.Tk()
        root.withdraw()     

        # 사용자 Excel file 경로 지정 요청
        time.sleep(1)
        print("== 기준 데이터가 될 엑셀 파일을 선택해주세요 ==")

        file_path = filedialog.askopenfilename()
        splitPath = file_path.split("/")[-1]
        print(f"\n- 선택된 엑셀 파일 : {splitPath}")

        # realPath = input("엑셀 파일 경로 입력 : ")
        # resultPath = realPath.replace("\\", "/")

        # print("resultPath : ", resultPath)

        # Excel Sheet 추출 및 사용자 시트 선택 제공
        # [로딩이벤트] 1. 로딩 이벤트를 위한 스레드 객체 생성
        stop_event = threading.Event()

        # [로딩이벤트] 2.로딩 이벤트 중 실행될 스레드
        loading_thread = threading.Thread(target=self.loading_animation, args=(stop_event,))
        loading_thread.start()

        # 파일 미선택시 오류 발생
        try :
            excelFile = pyxl.load_workbook(file_path)
        except Exception as e :
            print("오류가 발생했습니다.", e)

        # openpyxl 을 통한 엑셀 파일 및 시트 추출
        sheet = excelFile.sheetnames

        # [로딩이벤트] 3.로딩 이벤트 종료
        stop_event.set()
        loading_thread.join()

        # 키보드에 남아있는 버퍼 비우기
        while msvcrt.kbhit():
            msvcrt.getch()

        # 사용자 시트 선택 제공
        chosenSheet = int(input(f"\n::: 추출할 시트를 선택해주세요(번호입력) :::\n\n{sheet}\n : "))
        realSheet = sheet[chosenSheet-1]

        sheet = excelFile[realSheet]
        sheet_maxRow = sheet.max_row

        # 실제 엑셀 시트의 수입신고번호와
        # 사용자가 지정한 경로 내의 수입신고번호 폴더와 비교 후
        # 일치하면 폴더명 변경 진행
        self.matchNRename(chosenSheet, sheet, sheet_maxRow)


if __name__ == "__main__":
    app = FolderRenamer()
    
    # ESC 감지 스레드 별도 관리
    esc = threading.Thread(target=app.on_key_exit, daemon=True)
    esc.start()

    # 버전관리
    version = "ver 1.0"
    author = "Choi Yeon Woong"
    date = "2025.05.27"
    
    app.waitingForUser()
    app.on_key_title(version, author, date)
    app.on_key_start()

    print("모든 프로세스가 종료되었습니다.")








